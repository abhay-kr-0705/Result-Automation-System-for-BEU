from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_session import Session
import os
from datetime import datetime, timedelta
import re
import time
import io
from werkzeug.utils import secure_filename
from scraper import BEUResultScraper
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import xlsxwriter

app = Flask(__name__)
app.config['SECRET_KEY'] = 'beu-results-automation-2024'
app.config['SESSION_TYPE'] = 'filesystem'
Session(app)

# Branch codes mapping
BRANCH_CODES = {
    'Electronics Engineering (VLSI)': '159',
    'Computer Science & Engineering (CSE)': '105',
    'Electrical & Electronics Engineering (EEE)': '110',
    'Civil Engineering': '101',
    'Mining Engineering': '113',
    'Mechanical Engineering': '102'
}

# College code (constant)
COLLEGE_CODE = '124'

# College code to name mapping
COLLEGE_NAMES = {
    '124': 'Sher Shah Engineering College'
}

# Branch code to full name mapping
BRANCH_FULL_NAMES = {
    '159': 'Electronics Engineering (VLSI)',
    '105': 'Computer Science & Engineering (CSE)',
    '110': 'Electrical & Electronics Engineering (EEE)',
    '101': 'Civil Engineering',
    '113': 'Mining Engineering',
    '102': 'Mechanical Engineering'
}

# Login credentials
VALID_USERNAME = 'Result@SEC'
VALID_PASSWORD = 'SEC@Result12#'

class ResultProcessor:
    def __init__(self):
        pass
        
    def get_available_semesters(self, admission_year):
        """Get available semesters based on admission year and current date"""
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        # Calculate how many semesters should be available
        years_passed = current_year - admission_year
        
        # Assuming 2 semesters per year, and considering current month
        if current_month >= 7:  # After July, odd semester results are usually out
            available_semesters = (years_passed * 2) + 1
        else:  # Before July, even semester results from previous year
            available_semesters = years_passed * 2
        
        # Cap at 8 semesters maximum
        available_semesters = min(available_semesters, 8)
        
        return list(range(1, available_semesters + 1))
    
    def convert_to_dataframe(self, results_data):
        """Convert scraped results to data structure"""
        df_data = []
        
        for result in results_data:
            if result.get('error'):
                # Add error entries
                row = {
                    'Registration Number': result['registration_number'],
                    'Name': 'ERROR',
                    'Semester': result.get('semester', ''),
                    'Year': result.get('year', ''),
                    'Error': result['error'],
                    'SGPA': '',
                    'CGPA': '',
                    'Result': ''
                }
                df_data.append(row)
            else:
                # Add successful results
                row = {
                    'Registration Number': result['registration_number'],
                    'Name': result.get('name', ''),
                    'Semester': result.get('semester', ''),
                    'Year': result.get('year', ''),
                    'SGPA': result.get('sgpa', ''),
                    'CGPA': result.get('cgpa', ''),
                    'Result': result.get('result', ''),
                    'Error': ''
                }
                df_data.append(row)
        
        # Return the data structure directly (no pandas needed)
        return df_data
    
    def create_formatted_excel(self, results, filename, branch_code, admission_year, selected_semesters):
        """Create formatted Excel file with college header and multi-semester layout"""
        if not results:
            return None
        
        filepath = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        
        # Import required libraries for Excel formatting
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Get college and branch names
        college_name = COLLEGE_NAMES.get(COLLEGE_CODE, f"College Code {COLLEGE_CODE}")
        branch_name = BRANCH_FULL_NAMES.get(branch_code, f"Branch Code {branch_code}")
        
        # Header styles
        header_font = Font(name='Arial', size=14, bold=True)
        subheader_font = Font(name='Arial', size=12, bold=True)
        column_header_font = Font(name='Arial', size=10, bold=True)
        data_font = Font(name='Arial', size=9)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        semester_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 1: College Header
        ws.merge_cells('A1:Z1')  # Merge across many columns
        ws['A1'] = f"{COLLEGE_CODE} - {college_name}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_alignment
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        
        # Row 2: Course and Result Type
        ws.merge_cells('A2:Z2')
        ws['A2'] = f"{branch_name} Multi-Semester Results {admission_year}"
        ws['A2'].font = subheader_font
        ws['A2'].fill = subheader_fill
        ws['A2'].alignment = center_alignment
        
        # Row 3: Empty row for spacing
        ws.row_dimensions[3].height = 10
        
        # Row 4: Column headers
        current_row = 4
        
        # Column A: Registration No.
        ws['A4'] = "Registration No."
        ws['A4'].font = column_header_font
        ws['A4'].alignment = center_alignment
        
        # Column B: Name of Student
        ws['B4'] = "Name of Student"
        ws['B4'].font = column_header_font
        ws['B4'].alignment = center_alignment
        
        # Starting from Column C: Semester headers
        current_col = 3  # Column C
        semester_start_cols = {}
        
        # Organize results by registration number and semester
        student_data = {}
        all_subjects = set()
        
        for result in results:
            reg_num = result.get('registration_number', '')
            semester = result.get('semester', 0)
            
            if reg_num not in student_data:
                student_data[reg_num] = {
                    'name': result.get('name', result.get('student_name', '')),
                    'semesters': {}
                }
            
            student_data[reg_num]['semesters'][semester] = result
            
            # Collect all subjects for this semester
            if 'subjects' in result and result['subjects']:
                # Handle both dict and list formats
                if isinstance(result['subjects'], dict):
                    for subject_name in result['subjects'].keys():
                        all_subjects.add(f"S{semester}_{subject_name}")
                elif isinstance(result['subjects'], list):
                    for subject in result['subjects']:
                        if isinstance(subject, dict) and 'name' in subject:
                            all_subjects.add(f"S{semester}_{subject['name']}")
                        else:
                            all_subjects.add(f"S{semester}_{str(subject)}")
        
        # Create semester headers and sub-columns
        for semester in sorted(selected_semesters):
            semester_start_cols[semester] = current_col
            
            # Get subjects for this semester from results
            semester_subjects = []
            for result in results:
                if result.get('semester') == semester and 'subjects' in result and result['subjects']:
                    # Handle both dict and list formats
                    if isinstance(result['subjects'], dict):
                        for subject_name in result['subjects'].keys():
                            if subject_name not in semester_subjects:
                                semester_subjects.append(subject_name)
                    elif isinstance(result['subjects'], list):
                        for subject in result['subjects']:
                            if isinstance(subject, dict) and 'name' in subject:
                                if subject['name'] not in semester_subjects:
                                    semester_subjects.append(subject['name'])
                            else:
                                subject_str = str(subject)
                                if subject_str not in semester_subjects:
                                    semester_subjects.append(subject_str)
            
            # Calculate columns needed for this semester (SGPA + CGPA + subjects)
            cols_needed = 2 + len(semester_subjects)  # SGPA, CGPA, + subjects
            
            # Merge cells for semester header
            start_col_letter = get_column_letter(current_col)
            end_col_letter = get_column_letter(current_col + cols_needed - 1)
            ws.merge_cells(f'{start_col_letter}4:{end_col_letter}4')
            
            semester_cell = ws[f'{start_col_letter}4']
            semester_cell.value = f"SEMESTER {semester}"
            semester_cell.font = column_header_font
            semester_cell.fill = semester_fill
            semester_cell.alignment = center_alignment
            
            # Sub-headers for this semester (Row 5)
            sub_col = current_col
            
            # SGPA column
            ws[f'{get_column_letter(sub_col)}5'] = "SGPA"
            ws[f'{get_column_letter(sub_col)}5'].font = Font(name='Arial', size=9, bold=True)
            ws[f'{get_column_letter(sub_col)}5'].alignment = center_alignment
            sub_col += 1
            
            # CGPA column
            ws[f'{get_column_letter(sub_col)}5'] = "CGPA"
            ws[f'{get_column_letter(sub_col)}5'].font = Font(name='Arial', size=9, bold=True)
            ws[f'{get_column_letter(sub_col)}5'].alignment = center_alignment
            sub_col += 1
            
            # Subject columns
            for subject in semester_subjects:
                ws[f'{get_column_letter(sub_col)}5'] = subject
                ws[f'{get_column_letter(sub_col)}5'].font = Font(name='Arial', size=8, bold=True)
                ws[f'{get_column_letter(sub_col)}5'].alignment = center_alignment
                sub_col += 1
            
            current_col += cols_needed
        
        # Fill data rows starting from row 6
        data_row = 6
        for reg_num in sorted(student_data.keys()):
            student = student_data[reg_num]
            
            # Column A: Registration Number
            ws[f'A{data_row}'] = reg_num
            ws[f'A{data_row}'].font = data_font
            ws[f'A{data_row}'].alignment = center_alignment
            
            # Column B: Student Name
            ws[f'B{data_row}'] = student['name']
            ws[f'B{data_row}'].font = data_font
            
            # Fill semester data
            for semester in sorted(selected_semesters):
                if semester in student['semesters']:
                    result = student['semesters'][semester]
                    start_col = semester_start_cols[semester]
                    
                    # SGPA
                    ws[f'{get_column_letter(start_col)}{data_row}'] = result.get('sgpa', '')
                    ws[f'{get_column_letter(start_col)}{data_row}'].font = data_font
                    ws[f'{get_column_letter(start_col)}{data_row}'].alignment = center_alignment
                    
                    # CGPA
                    ws[f'{get_column_letter(start_col + 1)}{data_row}'] = result.get('cgpa', '')
                    ws[f'{get_column_letter(start_col + 1)}{data_row}'].font = data_font
                    ws[f'{get_column_letter(start_col + 1)}{data_row}'].alignment = center_alignment
                    
                    # Subject marks
                    if 'subjects' in result and result['subjects']:
                        subject_col = start_col + 2
                        if isinstance(result['subjects'], dict):
                            for subject_name, subject_data in result['subjects'].items():
                                marks = subject_data.get('marks', '') if isinstance(subject_data, dict) else subject_data
                                ws[f'{get_column_letter(subject_col)}{data_row}'] = marks
                                ws[f'{get_column_letter(subject_col)}{data_row}'].font = data_font
                                ws[f'{get_column_letter(subject_col)}{data_row}'].alignment = center_alignment
                                subject_col += 1
            
            data_row += 1
        
        # Apply borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=4, max_row=data_row-1, min_col=1, max_col=current_col-1):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col in range(1, current_col):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, data_row):
                cell_value = ws[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set minimum and maximum widths
            adjusted_width = min(max(max_length + 2, 10), 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(filepath)
        return filepath
    
    def save_to_excel(self, data, filename):
        """Save data to Excel file with formatting (legacy method)"""
        if not data:
            return None
        
        filepath = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        
        # Create workbook manually
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Get headers from first row
        if data:
            headers = list(data[0].keys())
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(header, ''))
        
        wb.save(filepath)
        return filepath
    
    def save_to_csv(self, results, filename):
        """Save results to CSV format"""
        if not results:
            return None
        
        filepath = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        
        # Collect all unique subject names
        all_subjects = set()
        for result in results:
            if 'subjects' in result and result['subjects']:
                if isinstance(result['subjects'], dict):
                    all_subjects.update(result['subjects'].keys())
        
        # Create CSV content manually
        csv_lines = []
        
        # Header row
        header = ['Registration Number', 'Name', 'Semester', 'Year', 'SGPA', 'CGPA', 'Result']
        for subject in sorted(all_subjects):
            header.extend([f'{subject}_Marks', f'{subject}_Grade'])
        csv_lines.append(','.join(header))
        
        # Data rows
        for result in results:
            row = [
                result.get('registration_number', ''),
                result.get('name', ''),
                str(result.get('semester', '')),
                str(result.get('year', '')),
                result.get('sgpa', ''),
                result.get('cgpa', ''),
                result.get('result', '')
            ]
            
            # Add subject data
            for subject in sorted(all_subjects):
                if 'subjects' in result and isinstance(result['subjects'], dict) and subject in result['subjects']:
                    subject_data = result['subjects'][subject]
                    if isinstance(subject_data, dict):
                        row.extend([subject_data.get('marks', ''), subject_data.get('grade', '')])
                    else:
                        row.extend([str(subject_data), ''])
                else:
                    row.extend(['', ''])
            
            csv_lines.append(','.join([f'"{str(cell)}"' for cell in row]))
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(csv_lines))
        
        return filepath

@app.route('/')
def login():
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/authenticate', methods=['POST'])
def authenticate():
    username = request.form.get('username')
    password = request.form.get('password')
    
    if username == VALID_USERNAME and password == VALID_PASSWORD:
        session['logged_in'] = True
        return redirect(url_for('dashboard'))
    else:
        return render_template('login.html', error='Invalid credentials')

@app.route('/dashboard')
def dashboard():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html', branch_codes=BRANCH_CODES)

@app.route('/get_available_semesters', methods=['POST'])
def get_available_semesters():
    admission_year = int(request.json.get('admission_year'))
    processor = ResultProcessor()
    available_semesters = processor.get_available_semesters(admission_year)
    return jsonify({'available_semesters': available_semesters})

@app.route('/scrape_results', methods=['POST'])
def scrape_results():
    if 'logged_in' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    scraper = None
    try:
        data = request.json
        admission_year = int(data.get('admission_year'))
        branch = data.get('branch')
        selected_semesters = data.get('semesters', [])
        start_reg = data.get('start_reg')
        end_reg = data.get('end_reg')
        publication_dates = data.get('publication_dates')
        export_format = data.get('format', 'excel')
        
        # Get branch code
        branch_code = BRANCH_CODES.get(branch)
        if not branch_code:
            return jsonify({'error': 'Invalid branch selected'}), 400
        
        # Initialize scraper
        scraper = BEUResultScraper()
        processor = ResultProcessor()
        
        # Generate registration numbers
        reg_numbers = scraper.generate_registration_numbers(
            admission_year, branch_code, start_reg, end_reg
        )
        
        # Calculate expected passout year if not provided
        if data.get('passout_year'):
            passout_year = int(data.get('passout_year'))
        else:
            # Calculate passout year (admission year + 4 for B.Tech)
            passout_year = admission_year + 4
        
        print(f"DEBUG: Admission Year: {admission_year}, Passout Year: {passout_year}")
        print(f"DEBUG: Selected Semesters: {selected_semesters}")
        
        # Get ALL available result links first (no filtering)
        available_links = scraper.get_available_result_links()
        
        print(f"DEBUG: Found {len(available_links)} total B.Tech result links:")
        for i, link in enumerate(available_links):
            print(f"  {i+1}. {link['text']}")
            print(f"     Semester: {link['semester']}, Year: {link['year']}")
            print(f"     Batch: {link['batch_session']}")
            print(f"     Batch Admission Year: {link.get('batch_admission_year', 'None')}")
            print(f"     Published: {link['published_date']}")
            print(f"     Special: {link['is_special']}")
            print()
        
        # Filter links using both admission year and passout year logic
        semester_links = []
        for semester in selected_semesters:
            print(f"DEBUG: Looking for Semester {semester}...")
            
            # Find all links for this semester
            semester_candidates = [link for link in available_links if link['semester'] == semester]
            print(f"  Found {len(semester_candidates)} links for semester {semester}")
            
            # Try different matching strategies
            matching_links = []
            
            # Strategy 1: Match by batch admission year
            for link in semester_candidates:
                if link.get('batch_admission_year') == admission_year:
                    matching_links.append(link)
                    print(f"  ✓ Matched by admission year: {link['text']} (Batch: {link['batch_session']})")
            
            # Strategy 2: If no matches, try matching by expected batch format
            if not matching_links:
                expected_batch = f"{admission_year}-{str(passout_year)[-2:]}"  # e.g., "2021-25"
                for link in semester_candidates:
                    if link['batch_session'] == expected_batch:
                        matching_links.append(link)
                        print(f"  ✓ Matched by batch format: {link['text']} (Batch: {link['batch_session']})")
            
            # Strategy 3: If still no matches, try partial matching
            if not matching_links:
                for link in semester_candidates:
                    if str(admission_year) in link['batch_session'] or str(passout_year) in link['batch_session']:
                        matching_links.append(link)
                        print(f"  ✓ Matched by partial year: {link['text']} (Batch: {link['batch_session']})")
            
            if matching_links:
                # Use the most recent published result
                try:
                    semester_link = max(matching_links, key=lambda x: datetime.strptime(x['published_date'], '%d-%m-%Y') if x['published_date'] else datetime.min)
                except:
                    semester_link = matching_links[0]
                
                semester_links.append(semester_link)
                print(f"  → Selected: {semester_link['text']} (Batch: {semester_link['batch_session']})")
            else:
                print(f"  ✗ No matches found for semester {semester}")
        
        if not semester_links:
            # Provide comprehensive error information
            available_info = {}
            for link in available_links:
                sem = link['semester']
                if sem not in available_info:
                    available_info[sem] = []
                available_info[sem].append(f"{link['batch_session']} ({link['published_date']})")
            
            error_msg = f"No matching semester results found.\n"
            error_msg += f"Requested: Admission {admission_year}, Passout {passout_year}, Semesters {selected_semesters}\n"
            error_msg += f"Available semesters and batches:\n"
            for sem, batches in sorted(available_info.items()):
                error_msg += f"  Semester {sem}: {', '.join(batches)}\n"
            
            return jsonify({'error': error_msg}), 404
        
        # Scrape results for all semesters with homepage return between each
        all_results = scraper.scrape_multiple_semesters(semester_links, reg_numbers, admission_year)
        
        if all_results:
            # Create file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if export_format.lower() == 'csv':
                filename = f'results_{branch_code}_{timestamp}.csv'
                # Convert to DataFrame for CSV
                df = processor.convert_to_dataframe(all_results)
                filepath = processor.save_to_csv(df, filename)
            else:
                filename = f'results_{branch_code}_{timestamp}.xlsx'
                # Use new formatted Excel method
                filepath = processor.create_formatted_excel(all_results, filename, branch_code, admission_year, selected_semesters)
            
            if filepath and os.path.exists(filepath):
                return jsonify({
                    'success': True,
                    'message': f'Successfully scraped {len(all_results)} results',
                    'download_url': f'/download/{filename}',
                    'total_results': len(all_results),
                    'successful_results': len([r for r in all_results if not r.get('error')]),
                    'failed_results': len([r for r in all_results if r.get('error')])
                })
            else:
                return jsonify({'error': 'Failed to create output file'}), 500
        else:
            return jsonify({'error': 'No results found for the specified criteria'}), 404
            
    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    finally:
        # Always close the scraper driver
        if scraper:
            scraper.close_driver()

@app.route('/download/<filename>')
def download_file(filename):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    filepath = os.path.join('temp', filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    else:
        return "File not found", 404

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
